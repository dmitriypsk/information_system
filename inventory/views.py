import openpyxl, datetime
from django.shortcuts import render, redirect, get_object_or_404, redirect
from .models import (Product, Category, Supplier, Transaction,
                     InventoryItem, CreditSale, CreditPayment,Supplier)
from .forms import (ProductForm, CategoryForm, TransactionForm,
                    CreditSaleForm, CreditPaymentForm, SupplierForm)
from django.contrib import messages
from django.http import HttpResponse
from datetime import timedelta, timezone, datetime, date
from django.urls import reverse
from django.db.models import F, Sum, FloatField, Q, Value, ExpressionWrapper, Case, When, IntegerField, Count, Prefetch, CharField
from django.db.models.functions import Coalesce
from inventory.models import Transaction, CreditSale, Product
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm


 # Лариса Kruglikova3 
 # Галя Vavilonpskov


def is_director(user):
    return user.groups.filter(name="Директор").exists()


def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                if user.groups.filter(name="Директор").exists():
                    request.session['is_director'] = True
                else:
                    request.session['is_director'] = False
                return redirect('product_list')  # замените 'product_list' на название представления главной страницы
            else:
                form.add_error(None, 'Неверный логин или пароль')

    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})  # замените 'login.html' на название шаблона для страницы входа
 # Создание оставшихся платежей автоматически 


@login_required
def product_list(request):
    def get_credit_sales_products():
        credit_sales = CreditSale.objects.all()
        credit_sales_products = set()
        for sale in credit_sales:
            for product in sale.products.all():
                credit_sales_products.add(product)
        return credit_sales_products

    all_products = Product.objects.all()
    credit_sales_products = get_credit_sales_products()

    if request.path == '/inventory/transactions/new/':
        available_products = [product for product in all_products if is_product_available(product) and product not in credit_sales_products]
    else:
        available_products = [product for product in all_products if is_product_available(product)]

    return render(request, 'inventory/product_list.html', {'inventory_items': available_products})


@login_required
@user_passes_test(is_director)
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers})


@login_required
@user_passes_test(is_director)
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'inventory/supplier_form.html', {'form': form, 'form_title': 'Добавить поставщика', 'button_label': 'Сохранить'})


@login_required
@user_passes_test(is_director)
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'inventory/supplier_form.html', {'form': form, 'form_title': 'Изменить поставщика', 'button_label': 'Сохранить'})


@login_required
@user_passes_test(is_director)
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_confirm_delete.html', {'supplier': supplier})


@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, user=request.user)
        if form.is_valid():
            product = form.save()
            InventoryItem.objects.create(product=product)
            return redirect('product_list')
    else:
        form = ProductForm(user=request.user)
    return render(request, 'inventory/product_form.html', {'form': form})


@login_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/product_form.html', {'form': form})


@login_required
@user_passes_test(is_director)
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'inventory/product_confirm_delete.html', {'product': product})


@login_required
@user_passes_test(is_director)
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'inventory/category_list.html', {'categories': categories})


@login_required
@user_passes_test(is_director)
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'inventory/category_form.html', {'form': form})


@login_required
@user_passes_test(is_director)
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'inventory/category_form.html', {'form': form})


@login_required
@user_passes_test(is_director)
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    category.delete()
    return redirect('category_list')


@login_required
def transaction_list(request):
    transactions = Transaction.objects.prefetch_related(
        Prefetch('products', queryset=Product.objects.all())
    ).annotate(
        total_discounted_price=Sum('products__discounted_price')
    ).order_by('-created_at')
    context = {
        'transactions': transactions,
    }
    return render(request, 'inventory/transaction_list.html', context)


@login_required
def transaction_create(request):
    if request.method == "POST":
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            products = form.cleaned_data['products']

            all_products_available = True

            for product in products:
                if transaction.transaction_type == Transaction.RETURN:
                    if product.status == Product.SOLD:
                        product.status = Product.IN_STOCK  # Изменили статус на IN_STOCK
                        product.save()
                    else:
                        all_products_available = False
                        messages.error(request, f'Товар {product.name} уже возвращен или недоступен для возврата.')
                else:
                    if product.status == Product.IN_STOCK:
                        product.status = Product.SOLD if transaction.transaction_type == Transaction.SALE else Product.WRITTEN_OFF
                        product.save()
                    else:
                        all_products_available = False
                        messages.error(request, f'Товар {product.name} недоступен для данной операции.')

            if all_products_available:
                transaction.save()
                for product in products:
                    transaction.products.add(product)  # Теперь добавляем продукты к связи "многие ко многим"
                return redirect('transaction_list')
            else:
                transaction.delete()

    else:
        form = TransactionForm()
    return render(request, 'inventory/transaction_form.html', {'form': form})


@login_required
@user_passes_test(is_director)
def report(request):
    credit_sales = CreditSale.objects.annotate(
        paid_amount=Coalesce(
            Sum('creditpayment__amount', filter=Q(creditpayment__paid=True)), Value(0), output_field=IntegerField()
        ),
    ).annotate(
        remaining_amount_value=ExpressionWrapper(F('total_price') - F('paid_amount'), output_field=IntegerField())
    )

    transaction_type_choices = Transaction._meta.get_field('transaction_type').choices
    transaction_type_dict = dict(transaction_type_choices)

    transactions = Transaction.objects.annotate(
        transaction_type_name=Case(
            *[When(transaction_type=key, then=Value(value)) for key, value in transaction_type_dict.items()],
            output_field=CharField()
        ),
    ).annotate(
        total_discounted_price=Sum('products__discounted_price')
    ).order_by('-created_at')

    in_stock_products = Product.objects.filter(status=Product.IN_STOCK)
    total_in_stock = len(in_stock_products)

    credit_sale_transactions = Transaction.objects.filter(transaction_type=Transaction.SALE, products__creditsale__products__isnull=False).values('products__name').annotate(
        total=Count('products__name', distinct=True)
    )

    context = {
        'transactions': transactions,
        'credit': credit_sales,
        'products_in_stock': in_stock_products,
        'credit_sale_transactions': credit_sale_transactions,
        'total_in_stock': total_in_stock,
    }

    return render(request, 'inventory/report.html', context)


@login_required
@user_passes_test(is_director)
def export_report_to_excel(request):
    def format_date_or_datetime(value):
        if isinstance(value, datetime):
            return value.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        else:
            return value
    # Создание книги Excel
    workbook = openpyxl.Workbook()

    # Лист "Инвентарь"
    inventory_data = InventoryItem.objects.select_related('product').all().order_by('product__name')
    inventory_worksheet = workbook.active
    inventory_worksheet.title = 'Инвентарь'

    # Заголовки столбцов для инвентаря
    inventory_columns = ['Товар', 'Артикул', 'Производитель', 'Категория', 'Цена закупки', 'Цена продажи', 'Скидка']
    for col_num, column_title in enumerate(inventory_columns, 1):
        cell = inventory_worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Заполнение строк данными инвентаря
    for row_num, item in enumerate(inventory_data, 2):
        product = item.product
        row = [
            product.name,
            product.sku,
            product.manufacturer,
            product.category.name,
            round(product.purchase_price),
            round(product.sale_price),
            round(product.discounted_price),
    ]
        for col_num, cell_value in enumerate(row, 1):
            cell = inventory_worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Настройка ширины столбцов для инвентаря
    for column_cells in inventory_worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        inventory_worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Лист "Транзакции и кредитные продажи"
    transactions_data = Transaction.objects.all().order_by('created_at')
    credit_sales_data = CreditSale.objects.all().order_by('sale_date')

    transactions_worksheet = workbook.create_sheet('Транзакции и кредитные продажи')

    # Заголовки столбцов для транзакций и кредитных продаж
    transactions_columns = ['Товар', 'Тип операции', 'Дата операции', 'Кредитная продажа']
    for col_num, column_title in enumerate(transactions_columns, 1):
        cell = transactions_worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Заполнение строк данными транзакций и кредитных продаж
    for row_num, transaction in enumerate(transactions_data, 2):
        products_names = ', '.join(transaction.products.values_list('name', flat=True))
        row = [
            products_names,
            transaction.transaction_type,
            format_date_or_datetime(transaction.created_at),
            'Нет',
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = transactions_worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        row_num += 1

    for credit_sale in credit_sales_data:
        products_names = ', '.join([product.name for product in credit_sale.products.all()])
        row = [
            products_names,
            'Продажа в кредит',
            format_date_or_datetime(credit_sale.sale_date),
            'Да',
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = transactions_worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        row_num += 1

    # Настройка ширины столбцов для транзакций и кредитных продаж

    for column_cells in transactions_worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        transactions_worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Лист "Неоплаченные кредиты"
    unpaid_credits_data = CreditPayment.objects.filter(paid=False).order_by('payment_date')

    unpaid_credits_worksheet = workbook.create_sheet('Неоплаченные кредиты')

    # Заголовки столбцов для неоплаченных кредитов
    unpaid_credits_columns = ['Товар', 'Дата платежа', 'Сумма платежа']
    for col_num, column_title in enumerate(unpaid_credits_columns, 1):
        cell = unpaid_credits_worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Заполнение строк данными неоплаченных кредитов
    for row_num, unpaid_credit in enumerate(unpaid_credits_data, 2):
        products_names = ', '.join([product.name for product in unpaid_credit.credit_sale.products.all()])
        row = [
            products_names,
            format_date_or_datetime(unpaid_credit.payment_date),
            unpaid_credit.amount,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = unpaid_credits_worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Настройка ширины столбцов для неоплаченных кредитов
    for column_cells in unpaid_credits_worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        unpaid_credits_worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Формирование HTTP-ответа с файлом Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'
    workbook.save(response)

    return response


@login_required
def transaction_return_create(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            products = form.cleaned_data['products']
            if transaction.transaction_type == Transaction.RETURN:
                for product in products:
                    if is_product_available(product):
                        product.status = Product.IN_STOCK
                        product.save()
                        transaction.products.add(product)
                    else:
                        messages.error(request, f'Товар {product.name} уже возвращен или продан в кредит.')
                transaction.save()
                return redirect('transaction_list')
            else:
                messages.error(request, 'Неверный тип операции.')
    else:
        form = TransactionForm(initial={'transaction_type': Transaction.RETURN})
    return render(request, 'inventory/transaction_form.html', {'form': form})


def is_product_available(product):
    return product.status == Product.IN_STOCK


@login_required
def credit_sales_list(request):
    credit_sales = CreditSale.objects.annotate(
        paid_amount=Coalesce(
            Sum('creditpayment__amount', filter=Q(creditpayment__paid=True)), Value(0), output_field=IntegerField()
        ),
    ).annotate(
        remaining_amount_value=ExpressionWrapper(F('total_price') - F('paid_amount'), output_field=IntegerField())
    )
    return render(request, 'inventory/credit_sales_list.html', {'credit_sales': credit_sales})


@login_required
def credit_sale_create(request):
    if request.method == 'POST':
        form = CreditSaleForm(request.POST)
        if form.is_valid():
            credit_sale = form.save()
            products = credit_sale.products.all()

            all_products_available = True

            for product in products:
                if is_product_available(product):
                    # Создайте новую запись в таблице Transaction
                    transaction = Transaction(transaction_type=Transaction.SALE)
                    transaction.save()
                    transaction.products.add(product)  # Добавьте продукт после сохранения транзакции

                    # Обновите статус продукта
                    product.status = Product.SOLD
                    product.save()
                else:
                    all_products_available = False
                    messages.error(request, f'Товар {product.name} недоступен для кредитной продажи.')

            if all_products_available:
                messages.success(request, 'Кредитная продажа успешно создана.')
                return redirect(reverse('credit_sales_list'))
            else:
                credit_sale.delete()

    else:
        form = CreditSaleForm()
    return render(request, 'inventory/credit_sale_form.html', {'form': form})


@login_required
def credit_payments_list(request, credit_sale_id):
    credit_sale = CreditSale.objects.get(pk=credit_sale_id)
    credit_payments = CreditPayment.objects.filter(credit_sale=credit_sale)
    return render(request, 'inventory/credit_payments_list.html', {'credit_sale': credit_sale, 'credit_payments': credit_payments})


@login_required
def credit_payment_create(request, credit_sale_id):
    credit_sale = get_object_or_404(CreditSale, id=credit_sale_id)

    if request.method == 'POST':
        form = CreditPaymentForm(request.POST)
        if form.is_valid():
            first_payment = form.save(commit=False)
            first_payment.credit_sale = credit_sale
            first_payment.save()

            # Создание оставшихся платежей автоматически
            remaining_months = credit_sale.credit_months - 1
            remaining_amount = credit_sale.total_price - first_payment.amount
            monthly_amount = remaining_amount / remaining_months
            payment_date = first_payment.payment_date

            for _ in range(remaining_months):
                payment_date += timedelta(days=30)  # прибавляем 30 дней к дате предыдущего платежа
                payment = CreditPayment(credit_sale=credit_sale, payment_date=payment_date, amount=monthly_amount)
                payment.save()

            return redirect(reverse('credit_payments_list', args=[credit_sale.id]))

    else:
        form = CreditPaymentForm()

    return render(request, 'inventory/credit_payment_form.html', {'form': form, 'credit_sale': credit_sale})


@login_required
def settled_credits(request):
    settled_credit_sales = CreditSale.objects.annotate(
        paid_amount=Coalesce(
            Sum('creditpayment__amount', filter=Q(creditpayment__paid=True)), Value(0), output_field=FloatField()
        ),
    ).annotate(
        remaining_amount_value=ExpressionWrapper(F('total_price') - F('paid_amount'), output_field=FloatField())
    ).filter(remaining_amount_value__lte=0)  # Добавьте этот фильтр

    return render(request, 'inventory/settled_credits.html', {'settled_credit_sales': settled_credit_sales})


@login_required
def unsettled_credits(request):
    unsettled_credit_sales = CreditSale.objects.filter(settled=False).annotate(
        paid_amount=Coalesce(
            Sum('creditpayment__amount', filter=Q(creditpayment__paid=True)), Value(0), output_field=FloatField()
        ),
    ).annotate(
        remaining_amount_value=ExpressionWrapper(F('total_price') - F('paid_amount'), output_field=FloatField())
    ).filter(remaining_amount_value__gt=0) 
    return render(request, 'inventory/unsettled_credits.html', {'unsettled_credit_sales': unsettled_credit_sales})


@login_required
def credit_return(request, credit_sale_id):
    credit_sale = get_object_or_404(CreditSale, id=credit_sale_id)
    products = credit_sale.products.all()

    if request.method == 'POST':
        # Изменение статуса кредитной продажи на "оплачено"
        credit_sale.settled = True
        credit_sale.save()  # Сохраните объект credit_sale перед созданием объекта Transaction

        # Создание новой транзакции возврата
        return_transaction = Transaction(transaction_type=Transaction.RETURN, total_price=credit_sale.total_price, credit_sale_id=credit_sale.id)
        return_transaction.save()

        for product in products:
            return_transaction.products.add(product)
            product.status = Product.IN_STOCK  # Обновление статуса продукта на 'in_stock'
            product.save()

        # Удаление кредитной продажи после сохранения объекта Transaction
        credit_sale.delete()

        return redirect('credit_sales')

    return render(request, 'inventory/credit_return.html', {'credit_sale': credit_sale})


@login_required
def credit_payment_toggle_paid(request, credit_payment_id):
    credit_payment = get_object_or_404(CreditPayment, pk=credit_payment_id)
    credit_payment.paid = not credit_payment.paid
    credit_payment.save()

    # Добавьте перенаправление обратно на страницу списка платежей
    return redirect('credit_payments_list', credit_sale_id=credit_payment.credit_sale.id)

